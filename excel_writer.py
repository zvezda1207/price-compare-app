from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def save_to_excel(df, path):

    df.to_excel(path, index=False)

    wb = load_workbook(path)
    ws = wb.active

    green_fill = PatternFill(
        start_color="00FF00", end_color="00FF00", fill_type="solid"
    )

    # начинаем со второй строки (первая — заголовки)
    for row in ws.iter_rows(min_row=2):

        # цены поставщиков начинаются с 3-й колонки
        supplier_cells = row[2:]

        prices = []

        for cell in supplier_cells:
            if cell.value is not None:
                prices.append(cell.value)

        if prices:
            min_price = min(prices)

            for cell in supplier_cells:
                if cell.value == min_price:
                    cell.fill = green_fill
    wb.save(path)


